###########################################################################
########################## DEVELOPED BY CAFTELLE ##########################
########################## DEVELOPED BY CAFTELLE ##########################
########################## DEVELOPED BY CAFTELLE ##########################
###########################################################################
################################# I E T ###################################
######################## V E R S I O N  2 . 0 . 0 #########################
###########################################################################
###########################################################################
################## Caftelle Created by Furkan ARINCI ######################
###########################################################################

# Kütüphaneler
from pathlib import Path
from playsound import playsound
import sys
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtGui import QImage
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import QTimer
import cv2 as cv
from PyQt5 import QtWidgets, uic
from pyzbar import pyzbar
from pyzbar.pyzbar import decode
import pytesseract
import xlsxwriter
import openpyxl
import datetime
import requests
import os
import mimetypes
from email.message import EmailMessage
import smtplib
import time
import cv2
from ui_main_window import *
import threading
from threading import Thread
import atexit
import faulthandler

#WindowsTesseract
"""""
try:
    pytesseract.pytesseract.tesseract_cmd = Path().cwd() / "tesseract.exe"
except:
    pass
"""""

class Ui(QWidget):

    def __init__(self):

        super(Ui, self).__init__()
        uic.loadUi('ui_main_window.ui', self)  # Sayfamızı yüklüyoruz.
        self.cap2 = cv2.VideoCapture(0)
        # read self.image in BGR format
        self.ret2, self.image2 = self.cap2.read()
        self.imgForSearch = self.image2
        self.cap2.release()

        # Bunlar sıralamada problem olmasın diye burada kayıtlı, tek tek yükseltme olayı bulunmakta.
        self.tesisline = 2
        self.iptalline = 2
        self.excelAc = False
        # Silme onaylanırsa yeni dosya açabilmek için onay

        self.dialog = askingPage()  # İkinci sayfamızın sınıfı, bunu çağırabilmek için kullanacağız.
        self.dialog.mySignal.connect(
            self.changeData)  # İkinci sayfamızdaki veri değiştiğinde ve butona tıklanıldığında burası çalışacak.
        self.scthread = Thread(target=self.showCamera)
        self.scthread.start()

        # dosya açma
        try:
            self.planWorkbook = openpyxl.load_workbook('Taranan_Is_Emirleri.xlsx')
            # Excel Stün ve Sekme Oluşturma
            self.planSheettesis = self.planWorkbook["TESIS"]
            self.planSheetiptal = self.planWorkbook["IPTAL"]
        except:
            self.mfg_print_and_show("Taranan_Is_Emirleri.xlsx dosyası oluşturulacaktır.")
            self.excelAc = True

        # Burada benim butonlarımın bağlantıları bulunmakta
        self.taramaYap.clicked.connect(self.taramaYapF)
        self.uiMailGonder.clicked.connect(self.uiMailGonderFonksiyon)
        self.uiMailGonder.setEnabled(False)
        self.myShow = True
        app.aboutToQuit.connect(self.closeThreads)

    def changeData(self, data):  # Veri değiştiğinde label adlı yazı bölgemize
        try:
            self.mfg_print_and_show('Dosyanın silinmesi hakkında...')
            self.dialog.show()
            if data == "delete":
                self.excelAc = True
                self.mfg_print_and_show("Dosya silindi!")
                tarananisemri2 = 'Taranan_Is_Emirleri.xlsx'
                os.remove(tarananisemri2)
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
            else:
                self.excelAc = False
                self.mfg_print_and_show("Dosya silinmedi!")
                alarmsound = Path().cwd() / "alarm.mp3"
                playsound(alarmsound)
        except Exception as e:
            self.excelAc = False
            print(e)
            self.mfg_print_and_show('Dosya zaten silinmiş.')
            alarmsound = Path().cwd() / "alarm.mp3"
            playsound(alarmsound)
            return False

    def taramaYapF(self):

        self.mfg_print_and_show('Belge taramayı başlattınız!')
        self.uiMailGonder.setEnabled(False)
        self.mfg_print_and_show('-')
        self.musteriNum.setText('-')
        self.hizmetNum.setText('-')
        self.isNum.setText('-')
        self.isTur.setText('-')
        self.awlthread = Thread(target=self.AllWithLove)
        self.awlthread.start()
        # Asıl fonksiyona yönlendirdim.

    def closeThreads(self):
        self.setShow(False)
        print("fdgd")

    # Kod karmaşasından kurtulmak için bunu yaptım.
    def mfg_print_and_show(self, data):
        self.bilgi.setText(data)
        print(f'{self.tesisline}-{self.iptalline}')
        print(data)

    def AllWithLove(self):

        print('')
        print(
            '################################################### DEVELOPED BY CAFTELLE ###################################################')
        print('')
        self.taramaYap.setEnabled(False)

        # FormDosyaAdıBelirleme
        bdtarih = datetime.datetime.now()
        yil = bdtarih.year
        ay = bdtarih.month
        gun = bdtarih.day
        saat = bdtarih.hour
        dakika = bdtarih.minute
        toplami = str(yil) + '_' + str(ay) + '_' + str(gun)
        dosyaadi = toplami + '_Tarihli_Is_Emirleri_Tutanagı.xlsx'
        dosyaadifinal = str(dosyaadi)

        tutanakdizinpath = str(os.getcwd())
        tutanakdosyasi = tutanakdizinpath + '/' + 'TutanakForm.xlsm'

        for root, dir, files in os.walk(tutanakdizinpath):

            if 'TutanakForm.xlsm' in files:
                self.mfg_print_and_show('Tutanak Formu Dosya içerisinde mevcut.\nİşleme devam ediyorum... ')
                break

            if not 'TutanakForm.xlsm' in files:
                self.mfg_print_and_show('Tutanak Formu bulunamadı\nHemen İndiriyorum... ')
                # FormDosyasıİndirme
                self.mfg_print_and_show('Tutanak Formu indiriliyor... ')
                resp = requests.get(
                    'https://www.dropbox.com/scl/fi/ydyx5isxb2szdhntamf85/TutanakForm.xlsm?dl=1&rlkey=g63u0w9uago9jajslnpejkjqd')

                with open('TutanakForm.xlsm', 'wb') as output:
                    output.write(resp.content)
                    self.mfg_print_and_show('İndirme Tamamlandı.')
                    textsound = Path().cwd() / "text.mp3"
                    playsound(textsound)

                break

        # Excel Dosyası Oluşturma
        if self.excelAc == True:
            try:
                self.planWorkbook1 = xlsxwriter.Workbook('Taranan_Is_Emirleri.xlsx')
                self.planSheettesis12 = self.planWorkbook1.add_worksheet("TESIS")
                self.planSheetiptal12 = self.planWorkbook1.add_worksheet("IPTAL")
                self.planWorkbook1.close()
                self.planWorkbook = openpyxl.load_workbook('Taranan_Is_Emirleri.xlsx')

                # Excel Stün ve Sekme Oluşturma
                self.planSheettesis = self.planWorkbook["TESIS"]
                self.planSheetiptal = self.planWorkbook["IPTAL"]

                self.planSheettesis['A1'] = 'Hizmet Numarası'
                self.planSheettesis['B1'] = 'Müşteri Numarası'
                self.planSheettesis['C1'] = 'İş Emri Numarası'
                self.planSheettesis['C1'] = 'İş Emri Numarası'
                self.planSheettesis['D1'] = 'Hizmet Türü'
                self.planSheettesis['E1'] = 'İş Emri Tipi'
                self.planSheettesis['F1'] = 'Tarih'

                self.planSheetiptal['A1'] = 'Hizmet Numarası'
                self.planSheetiptal['B1'] = 'Müşteri Numarası'
                self.planSheetiptal['C1'] = 'İş Emri Numarası'
                self.planSheetiptal['D1'] = 'Hizmet Türü'
                self.planSheetiptal['E1'] = 'İş Emri Tipi'
                self.planSheetiptal['F1'] = 'Tarih'
            except:
                self.mfg_print_and_show("Taranan_Is_Emirleri.xlsx dosyası oluşturulacaktır.")
        else:
            try:
                # kaçıncı satırın boş olduğunu öğreneceğim
                # ilk başta dosyamı açıyorum
                self.planWorkbook = openpyxl.load_workbook('Taranan_Is_Emirleri.xlsx')
                # Excel Stün ve Sekme Oluşturma
                self.planSheettesis = self.planWorkbook["TESIS"]
                self.planSheetiptal = self.planWorkbook["IPTAL"]

                self.planSheettesis['A1'] = 'Hizmet Numarası'
                self.planSheettesis['B1'] = 'Müşteri Numarası'
                self.planSheettesis['C1'] = 'İş Emri Numarası'
                self.planSheettesis['C1'] = 'İş Emri Numarası'
                self.planSheettesis['D1'] = 'Hizmet Türü'
                self.planSheettesis['E1'] = 'İş Emri Tipi'
                self.planSheettesis['F1'] = 'Tarih'

                self.planSheetiptal['A1'] = 'Hizmet Numarası'
                self.planSheetiptal['B1'] = 'Müşteri Numarası'
                self.planSheetiptal['C1'] = 'İş Emri Numarası'
                self.planSheetiptal['D1'] = 'Hizmet Türü'
                self.planSheetiptal['E1'] = 'İş Emri Tipi'
                self.planSheetiptal['F1'] = 'Tarih'

                i, j = 2, 2

                # sonra içinde geziyorum boş yer bulana kadar
                while self.getShow() and self.planSheettesis[f"A{i}"].value != None:
                    # time.sleep(1)
                    print(self.planSheettesis[f"A{i}"].value)
                    i += 1

                while self.getShow() and self.planSheetiptal[f"A{j}"].value != None:
                    # time.sleep(1)
                    print(self.planSheetiptal[f"A{j}"].value)
                    j += 1

                # boş yeri bulduğumda değer olarak yazdırıyorum.
                self.tesisline = i
                print(self.tesisline)
                self.iptalline = j
                print(self.iptalline)

            except:
                self.mfg_print_and_show("Taranan_Is_Emirleri.xlsx dosyası oluşturulacaktır.")

        # Değerlerin Sıfırlanması
        textstart = False
        savestart = True
        qrstart = True
        gerekli = False
        musterinofinal = '(     )'
        hizmetnofinal = '(     )'
        isemrinofinal = '1'
        isemrituru = '1'
        iptalturu = '(     )'
        isemriturufinal = '(     )'
        musterinoindex = 0
        isemrinoindex = 0
        qrhizmetno = 0
        isemrinoindex = 0
        isemrituruindex = 0
        # self.cap1 = cv2.VideoCapture(0)
        self.image = self.getImage()
        # convert self.image to RGB format
        self.image = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)

        while self.getShow() and qrstart:
            # Değerleri Sıfırlama
            musterinoindex = 0
            isemrinoindex = 0
            qrhizmetno = 0
            isemrinoindex = 0
            isemrituruindex = 0

            # Kamera'dan Aldığı Veriyi Okuma
            # get self.image infos
            self.image = self.getImage()
            self.height, width, channel = self.image.shape
            self.step = channel * width
            # create QImage from self.image
            qImg = QImage(self.image.data, width, self.height, self.step, QImage.Format_RGB888)
            # self.image_label.setPixmap(QPixmap.fromImage(qImg))

            font = cv.FONT_ITALIC
            decodedObjects = pyzbar.decode(self.image)

            for obj in decodedObjects:
                qrtemiz2 = obj.data.decode('utf-8')
                cv.putText(self.image, str(qrtemiz2), (200, 200), font, 1,
                           (0, 255, 160), 2)
                qrsound = Path().cwd() / "qr.mp3"
                playsound(qrsound)

            self.mfg_print_and_show(
                'Müşteri No, Hizmet No, İş Emri No Taranıyor...\nİş Emri Türü bir sonraki aşamada taranacak.')
            # cv.imshow("QR Tarama", self.image)
            # cv.waitKey(1)

            for qrcodee in decode(self.image):

                # Kamera'dan Alınan Verideki Yazıları Okuma
                self.mfg_print_and_show('QR Okundu ve Analiz Ediliyor...')
                qrtemiz = qrcodee.data.decode('utf-8')
                qrlist = qrtemiz.split("|")
                qrlistno = len(qrlist)

                # MÜSTERİ NO AYIKLAMA QR
                musterinoindex = [datano for datano in range(qrlistno) if qrlist[datano].startswith('M')]
                qrmusterino = qrlist[musterinoindex[0]]
                musterinofinal = qrmusterino.replace("M", "")
                self.mfg_print_and_show('Taranan QR Code içinden Müşteri Numarası ayıklanıyor...')

                # HİZMET NO AYIKLAMA QR
                hizmetnoindex = [datano for datano in range(qrlistno) if qrlist[datano].startswith('H')]
                qrhizmetno = qrlist[hizmetnoindex[0]]
                hizmetnofinal = qrhizmetno.replace("H", "")
                self.mfg_print_and_show('Taranan QR Code içinden Hizmet Numarası ayıklanıyor...')

                try:
                    # İş Emri Türü AYIKLAMA QR Eğer Eklenirse
                    isemrituruindex = [datano for datano in range(qrlistno) if qrlist[datano].startswith('IT')]
                    isemrituru1 = qrlist[isemrituruindex[0]]
                    isemriturufinal = isemrituru1.replace("IT", "")

                    replace_chars = [('ı', 'i'), ('İ', 'I'), ('ü', 'u'), ('Ü', 'U'), ('ö', 'o'), ('Ö', 'O'), ('ç', 'c'),
                                     ('Ç', 'C'),
                                     ('ş', 's'), ('Ş', 'S'), ('ğ', 'g'), ('Ğ', 'G')]

                    for search, replace in replace_chars:
                        isemriturufinal = isemriturufinal.replace(search, replace)
                        isemriturufinal = isemriturufinal
                        isemriturufinal = isemriturufinal.upper()
                        isemriturufinal = isemriturufinal.strip()
                        text2 = isemriturufinal

                except Exception as e:
                    print(e)
                    self.mfg_print_and_show('Hata algılandı!')

                # İŞ EMRİ NO AYIKLAMA QR
                isemrinoindex = [datano for datano in range(qrlistno) if qrlist[datano].startswith('I')]
                isemrinoindex2 = len(isemrinoindex)
                if isemrinoindex2 > 0:
                    qrisemrino = qrlist[isemrinoindex[0]]
                    isemrinofinal = qrisemrino.replace("I", "")
                    self.mfg_print_and_show('Taranan QR Code içinden İş Emri Numarası ayıklanıyor...')

                # Müşteri Numarasının ve Hizmet Numarasının Alındığını Doğrulama
                if musterinofinal != '(     )' and hizmetnofinal != '(     )':
                    qrstart = False

        if musterinofinal != '(     )' and hizmetnofinal != '(     )':
            cv.destroyAllWindows()
            textstart = True

        if isemriturufinal != '(     )':
            textstart = False
            gerekli = True

        # YAZI TARAMA
        # self.cap1 = cv2.VideoCapture(0)
        self.image = self.getImage()
        # convert self.image to RGB format
        self.image = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)

        while self.getShow() and textstart:

            # Kamera'dan Yazıları Okuma
            self.image = self.getImage()
            self.height, width, channel = self.image.shape
            self.step = channel * width
            # create QImage from self.image
            qImg = QImage(self.image.data, width, self.height, self.step, QImage.Format_RGB888)
            # self.image_label.setPixmap(QPixmap.fromImage(qImg))
            self.mfg_print_and_show(
                'İş Emri Türü: Taranıyor... Lütfen kameraya gösteriniz.')
            self.musteriNum.setText(musterinofinal)
            self.hizmetNum.setText(hizmetnofinal)
            self.isNum.setText(isemrinofinal)
            self.isTur.setText("Taranıyor...")

            text2 = pytesseract.image_to_string(self.image)
            boxes = pytesseract.image_to_boxes(self.image)

            for b in boxes.splitlines():
                b = b.split(' ')
                self.image = cv.rectangle(self.image, (int(b[1]), self.height - int(b[2])),
                                          (int(b[3]), self.height - int(b[4])), (0, 255, 160), 1)

            replace_chars = [('ı', 'i'), ('İ', 'I'), ('ü', 'u'), ('Ü', 'U'), ('ö', 'o'), ('Ö', 'O'), ('ç', 'c'),
                             ('Ç', 'C'),
                             ('ş', 's'), ('Ş', 'S'), ('ğ', 'g'), ('Ğ', 'G')]
            for search, replace in replace_chars:
                text2 = text2.replace(search, replace)
                text2 = text2.upper()
                text2 = text2.strip()
                break

            # Okunan Yazıları Tanıma ve Türüne Göre Ayıklama
            for self.image in text2:

                if 'OKUNAMIYOR' in text2:
                    iptalturu = 'OKUNAMADI'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show(
                        'İş Emri Türü: ' + 'OKUNMUYOR.')
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)

                    break

                if 'NUMARA TASIMA' in text2:
                    iptalturu = 'NUMARA TAŞIMALI YENİ ABONELİK'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KABLOSES IPTAL' in text2:
                    iptalturu = 'KABLOSES İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'ABONE ISTEGI ILE KABLOSES IPTALI' in text2:
                    iptalturu = 'ABONE ISTEGI ILE KABLOSES IPTALI'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'VERASETEN' and 'VERASETEN IPTAL' and 'VERASET' in text2:
                    iptalturu = 'VERASETEN İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'ABONE ISTEGI ILE IPTAL' in text2:
                    iptalturu = 'ABONELİK İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'ABONELIK IPTAL' in text2:
                    iptalturu = 'ABONELİK İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TARIFE DEGISIMI' in text2:
                    iptalturu = 'TARİFE DEĞİŞİMİ'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KAMPANYAYA GECIS' in text2:
                    iptalturu = 'TARİFE DEĞİŞİMİ'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KIRALAMA IPTAL' and 'CIHAZ KIRALAMA IPTAL' in text2:
                    iptalturu = 'CİHAZ KİRALAMA İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KIRALAMA SIPARIS' and 'CIHAZ KIRALAMA SIPARIS' in text2:
                    iptalturu = 'CİHAZ KİRALAMA SİPARİŞ'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLU ABONELIK DEVIR ALMA' and 'ABONELIK DEVIR ALMA' in text2:
                    iptalturu = 'TAAHHÜTLÜ ABONELİK DEVİR ALMA'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLU ABONELIK DEVIR' and 'ABONELIK DEVIR' in text2:
                    iptalturu = 'TAAHHÜTLÜ ABONELİK DEVİR'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'YENI ABONELIK' in text2:
                    iptalturu = 'TAAHHÜTLÜ YENİ ABONELİK'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'NAKIL GELEN' in text2:
                    iptalturu = 'TAAHHÜTLÜ NAKİL GELEN'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'ASKIYA ALMA' in text2:
                    iptalturu = 'HİZMETİ ASKIYA ALMA'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'HIZMETI ASKIYA ALMA' in text2:
                    iptalturu = 'HİZMETİ ASKIYA ALMA'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'CIHAZ IADE' and 'CIHAZ IADE FORMU' and 'IADE FORMU' in text2:
                    iptalturu = 'CİHAZ İADE FORMU'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KABLONET IPTAL BAŞVURU FORMU' and 'KABLOTV IPTAL BASVURU FORMU' and 'IPTAL BASVURU' in text2:
                    iptalturu = 'KABLONET İPTAL FORMU'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break


        while self.getShow() and gerekli:
            replace_chars = [('ı', 'i'), ('İ', 'I'), ('ü', 'u'), ('Ü', 'U'), ('ö', 'o'), ('Ö', 'O'), ('ç', 'c'),
                             ('Ç', 'C'),
                             ('ş', 's'), ('Ş', 'S'), ('ğ', 'g'), ('Ğ', 'G')]
            for search, replace in replace_chars:
                text2 = text2.replace(search, replace)
                text2 = text2
                text2 = text2.upper()
                text2 = text2.strip()
                break

            # Okunan Yazıları Tanıma ve Türüne Göre Ayıklama
            for self.image in text2:

                if 'OKUNAMIYOR' in text2:
                    iptalturu = 'OKUNAMADI'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'NUMARATASIMA' in text2:
                    iptalturu = 'NUMARA TAŞIMALI YENİ ABONELİK'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KABLOSESIPTAL' in text2:
                    iptalturu = 'KABLOSES İPTAL'
                    textstart = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'VERASETENIPTAL' in text2:
                    iptalturu = 'VERASETEN İPTAL'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'ABONELIKIPTAL' in text2:
                    iptalturu = 'ABONELİK İPTAL'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TARIFEDEGISIMI' in text2:
                    iptalturu = 'TARİFE DEĞİŞİMİ'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'CIHAZKIRALAMAIPTAL' in text2:
                    iptalturu = 'CİHAZ KİRALAMA İPTAL'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'CIHAZKIRALAMASIPARIS' in text2:
                    iptalturu = 'CİHAZ KİRALAMA SİPARİŞ'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLUABONELIKDEVIRALMA' in text2:
                    iptalturu = 'TAAHHÜTLÜ ABONELİK DEVİR ALMA'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLUABONELIKDEVIR' in text2:
                    iptalturu = 'TAAHHÜTLÜ ABONELİK DEVİR'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLUYENIABONELIK' in text2:
                    iptalturu = 'TAAHHÜTLÜ YENİ ABONELİK'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'TAAHHUTLU ABONELIK NAKIL GELEN' and 'TAHHUTLUNAKILGELEN' in text2:
                    iptalturu = 'TAAHHÜTLÜ NAKİL GELEN'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'HIZMETIASKIYAALMA' in text2:
                    iptalturu = 'HİZMETİ ASKIYA ALMA'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'CIHAZIADEFORMU' in text2:
                    iptalturu = 'CİHAZ İADE FORMU'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

                if 'KABLONETIPTALBASVURUFORMU' in text2:
                    iptalturu = 'KABLONET İPTAL FORMU'
                    gerekli = False
                    print('')
                    print(
                        '################################################### DEVELOPED BY CAFTELLE ###################################################')
                    print('')
                    self.mfg_print_and_show('İptal edildi: ' + iptalturu)
                    self.musteriNum.setText(musterinofinal)
                    self.hizmetNum.setText(hizmetnofinal)
                    self.isNum.setText(isemrinofinal)
                    self.isTur.setText(iptalturu)
                    break

        # Kamera'dan Okunan Yazının ayrıştırılıp Uygun Yere Atanması Kontrolü
        if iptalturu != '(     )':
            savestart = True

        # AYRIŞTIRILMIŞ VERİLERİ EXCELE KAYDETME

        while self.getShow() and savestart:

            # Taranan ve Ayrıştırılan Yazının Excel Tablosu Üzerinde Bulunan Sekmelerden Uygun Olana Atanması

            if 'OKUNAMADI' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = 'OKUNAMADI'
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı. Manuel Düzeltme Gerekiyor.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'NUMARA TAŞIMALI YENİ ABONELİK' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'KABLO SES'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'KABLOSES İPTAL' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'KABLO SES'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'TARİFE DEĞİŞİMİ' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')
                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'TAAHHÜTLÜ YENİ ABONELİK' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'TAAHHÜTLÜ ABONELİK DEVİR' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)

                break

            if 'HİZMETİ ASKIYA ALMA' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'TAAHHÜTLÜ ABONELİK DEVİR ALMA' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)

                break

            if 'CİHAZ KİRALAMA SİPARİŞ' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break
            if 'CİHAZ KİRALAMA İPTAL' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'TAAHHÜTLÜ NAKİL GELEN' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.tesisline)
                Bi = 'B' + str(self.tesisline)
                Ci = 'C' + str(self.tesisline)
                Di = 'D' + str(self.tesisline)
                Ei = 'E' + str(self.tesisline)
                Fi = 'F' + str(self.tesisline)

                self.planSheettesis[Ai] = hizmetnofinal
                self.planSheettesis[Bi] = musterinofinal
                self.planSheettesis[Ci] = isemrinofinal
                self.planSheettesis[Di] = iptalturu
                self.planSheettesis[Ei] = 'ANALOG KABLO TV'
                self.planSheettesis[Fi] = tarihcikti

                self.tesisline = self.tesisline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'ABONELİK İPTAL' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.iptalline)
                Bi = 'B' + str(self.iptalline)
                Ci = 'C' + str(self.iptalline)
                Di = 'D' + str(self.iptalline)
                Ei = 'E' + str(self.iptalline)
                Fi = 'F' + str(self.iptalline)

                self.planSheetiptal[Ai] = hizmetnofinal
                self.planSheetiptal[Bi] = musterinofinal
                self.planSheetiptal[Ci] = isemrinofinal
                self.planSheetiptal[Di] = iptalturu
                self.planSheetiptal[Ei] = 'ANALOG KABLO TV'
                self.planSheetiptal[Fi] = tarihcikti

                self.iptalline = self.iptalline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'CİHAZ İADE FORMU' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.iptalline)
                Bi = 'B' + str(self.iptalline)
                Ci = 'C' + str(self.iptalline)
                Di = 'D' + str(self.iptalline)
                Ei = 'E' + str(self.iptalline)
                Fi = 'F' + str(self.iptalline)

                self.planSheetiptal[Ai] = hizmetnofinal
                self.planSheetiptal[Bi] = musterinofinal
                self.planSheetiptal[Ci] = isemrinofinal
                self.planSheetiptal[Di] = iptalturu
                self.planSheetiptal[Ei] = 'ANALOG KABLO TV'
                self.planSheetiptal[Fi] = tarihcikti

                self.iptalline = self.iptalline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'VERASETEN İPTAL' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.iptalline)
                Bi = 'B' + str(self.iptalline)
                Ci = 'C' + str(self.iptalline)
                Di = 'D' + str(self.iptalline)
                Ei = 'E' + str(self.iptalline)
                Fi = 'F' + str(self.iptalline)

                self.planSheetiptal[Ai] = hizmetnofinal
                self.planSheetiptal[Bi] = musterinofinal
                self.planSheetiptal[Ci] = isemrinofinal
                self.planSheetiptal[Di] = iptalturu
                self.planSheetiptal[Ei] = 'ANALOG KABLO TV'
                self.planSheetiptal[Fi] = tarihcikti

                self.iptalline = self.iptalline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')

                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)
                break

            if 'KABLONET İPTAL FORMU' in iptalturu:
                tarih1 = datetime.datetime.now()
                yil1 = tarih1.year
                ay1 = tarih1.month
                gun1 = tarih1.day
                saat1 = tarih1.hour
                dakika1 = tarih1.minute
                toplami1 = str(yil1) + '/' + str(ay1) + '/' + str(gun1) + ' - ' + str(
                    saat1) + ':' + str(dakika1)
                tarihcikti = str(toplami1)

                Ai = 'A' + str(self.iptalline)
                Bi = 'B' + str(self.iptalline)
                Ci = 'C' + str(self.iptalline)
                Di = 'D' + str(self.iptalline)
                Ei = 'E' + str(self.iptalline)
                Fi = 'F' + str(self.iptalline)

                self.planSheetiptal[Ai] = hizmetnofinal
                self.planSheetiptal[Bi] = musterinofinal
                self.planSheetiptal[Ci] = isemrinofinal
                self.planSheetiptal[Di] = iptalturu
                self.planSheetiptal[Ei] = 'ANALOG KABLO TV'
                self.planSheetiptal[Fi] = tarihcikti

                self.iptalline = self.iptalline + 1
                savestart = False
                self.mfg_print_and_show('Başarıyla Aktarıldı.')
                self.planWorkbook.save('Taranan_Is_Emirleri.xlsx')
                textsound = Path().cwd() / "text.mp3"
                playsound(textsound)

                break

        print('')
        print(
            '################################################### DEVELOPED BY CAFTELLE ###################################################')
        print('')
        self.mfg_print_and_show(
            "Belge tarama bittiyse mail ile göndermek için\nGönder tuşuna basınız. Devam etmek istiyorsanız 'Tarama Yap'\ntuşuna basınız.")
        self.uiMailGonder.setEnabled(True)
        self.excelAc = False
        self.taramaYap.setEnabled(True)

    def dosyasilme(self):
        self.mfg_print_and_show('Dosyanın silinmesi hakkında...')
        self.dialog.show()

    def uiMailGonderFonksiyon(self):
        self.uiMailGonder.setEnabled(False)
        self.taramaYap.setEnabled(False)
        try:
            self.mfg_print_and_show('Mail gönderim tuşuna tıklandı!')
            mailOrName = self.mail.text()
            mailEk = '@turksat.com.tr'
            mail = ''

            if '@' in mailOrName:
                mail = mailOrName
            else:
                mail = mailOrName + mailEk

            recipient = mail
            bdtarih = datetime.datetime.now()
            yil = bdtarih.year
            ay = bdtarih.month
            gun = bdtarih.day
            saat = bdtarih.hour
            dakika = bdtarih.minute
            toplami = str(yil) + '_' + str(ay) + '_' + str(gun)
            dosyaadi = toplami + '_Tarihli_Is_Emirleri_Tutanagı.xlsm'
            dosyaadifinal = str(dosyaadi)

            # Yazılan Dosyayı Arama
            tutanakdizinpath2 = str(os.getcwd())
            tutanakdosyasi2 = tutanakdizinpath2 + '/' + 'TutanakForm.xlsm'

            if os.path.isfile(tutanakdosyasi2):

                tttarih = datetime.datetime.now()
                ttyil = tttarih.year
                ttay = tttarih.month
                ttgun = tttarih.day
                ttsaat = tttarih.hour
                ttdakika = tttarih.minute
                tttoplami = str(ttyil) + '_' + str(ttay) + '_' + str(ttgun)
                ttdosyaadi = tttoplami + '_Tarihli_Is_Emirleri_Tutanagı.xlsm'
                ttdosyaadifinal = str(ttdosyaadi)
                tarananisemri = 'Taranan_Is_Emirleri.xlsx'

                print('Gönderilecek Dosya Bulundu. Mail göndermeye hazırlanıyorum.')
                self.mfg_print_and_show('Gönderilecek Dosya Bulundu. Mail göndermeye hazırlanıyorum.')
                mail_server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
                mail_server.login("********@gmail.com", '*********')
                message = EmailMessage()
                sender = "developed.by.caftelle@gmail.com"
                recipient = mail
                message['From'] = 'Caftelle Software'
                message['To'] = recipient
                message['Subject'] = toplami + ' Tarihli İş Emirleri Tutanağı'
                body = 'Merhabalar, \n\n' + toplami + ' Tarihli İş Emirleri Ektedir.\n\nİyi Calismalar. \n \n \n | Developed by Caftelle  \n | Caftelle Created by Furkan ARINCI'
                message.set_content(body)
                mime_type, _ = mimetypes.guess_type(ttdosyaadifinal)
                mime_type, mime_subtype = mime_type.split('/')

                with open(tutanakdosyasi2, 'rb') as file:
                    message.add_attachment(file.read(), maintype=mime_type, subtype=mime_subtype,
                                           filename=tttoplami + '_Tarihli_Is_Emirleri_Tutanagı.xlsm')
                    print('Taranan Tutanak Formunu maile ekledim... ')
                    self.mfg_print_and_show('Taranan Tutanak Formunu maile ekledim... ')

                with open(tarananisemri, 'rb') as file:
                    self.mfg_print_and_show('Taranan Is Emirleri Formunu maile ekledim... ')
                    message.add_attachment(file.read(), maintype=mime_type, subtype=mime_subtype,
                                           filename='Taranan_Is_Emirleri.xlsx')

                mail_server.send_message(message)
                mail_server.quit()

                print('Gönderilen Mail Adresi: ' + recipient + '\nMail başarı ile gönderildi.')
                self.mfg_print_and_show('Gönderilen Mail Adresi: ' + recipient + '\nMail başarı ile gönderildi.')
                self.dosyasilme()
                mailsound = Path().cwd() / "mailgonder.mp3"
                playsound(mailsound)

            else:
                print('Dosya bulunamadığı için mail gönderilemedi.')
                self.mfg_print_and_show('Dosya bulunamadığı için mail gönderilemedi.')
                alarmsound = Path().cwd() / "alarm.mp3"
                playsound(alarmsound)

        except Exception as e:
            print(e)
            print('Mail Adresini veya Kullanıcı Adı yanlış olduğu için mail gönderilemedi.')
            self.mfg_print_and_show('Mail Adresini veya Kullanıcı Adı yanlış olduğu için mail gönderilemedi.')
            alarmsound = Path().cwd() / "alarm.mp3"
            playsound(alarmsound)

        self.uiMailGonder.setEnabled(True)
        self.taramaYap.setEnabled(True)

    # getter method
    def getImage(self):
        return self.imgForSearch

    # setter method
    def setImage(self, x):
        self.imgForSearch = x

    # getter method
    def getShow(self):
        return self.myShow

    # setter method
    def setShow(self, x):
        self.myShow = x

    # Burada kamera ekrana yansıtma yapacak
    def showCamera(self):
        self.cap2 = cv2.VideoCapture(0)
        # read self.image in BGR format
        self.ret2, self.image2 = self.cap2.read()
        self.setImage(self.image2)
        # convert self.image to RGB format
        self.image2 = cv2.cvtColor(self.image2, cv2.COLOR_BGR2RGB)
        # get self.image infos
        self.height2, width2, channel2 = self.image2.shape
        self.step2 = channel2 * width2
        # create QImage from self.image
        while self.getShow():
            time.sleep(0.03)
            self.ret2, self.image2 = self.cap2.read()
            if self.ret2:
                self.setImage(self.image2)
                self.image2 = cv2.cvtColor(self.image2, cv2.COLOR_BGR2RGB)
                qImg2 = QImage(self.image2.data, width2, self.height2, self.step2, QImage.Format_RGB888)
                # show self.image in img_label
                self.image_label.setPixmap(QPixmap.fromImage(qImg2))


class askingPage(QtWidgets.QDialog):  # ikinci sayfanin sinifi burada bulunmaktadir.

    mySignal = pyqtSignal(str)

    def __init__(self):
        super(askingPage, self).__init__()
        uic.loadUi('dosya_silme_onay.ui', self)  # İkinci sayfamızı buradan açıyoruz.

        self.deleteButton.clicked.connect(self.deleteButtonFunc)  # Butona tıkladığımızda

    def deleteButtonFunc(self):
        data = "delete"
        self.mySignal.emit(data)  # Şimdi verimizle beraber sinyal gönderiyoruz alıcıya.
        self.deleteButton.setText("Silindi")
        textsound = Path().cwd() / "text.mp3"
        playsound(textsound)
        self.close()


if __name__ == '__main__':
    # Bu kısım ilk başta çalışmaktadır.
    faulthandler.enable()
    app = QtWidgets.QApplication(sys.argv)
    window = Ui()
    window.show()
    app.exec_()


###########################################################################
###########################################################################
########################## DEVELOPED BY CAFTELLE ##########################
########################## DEVELOPED BY CAFTELLE ##########################
########################## DEVELOPED BY CAFTELLE ##########################
###########################################################################
################################# I E T ###################################
######################## V E R S I O N  2 . 0 . 0 #########################
###########################################################################
###########################################################################
################## Caftelle Created by Furkan ARINCI ######################
###########################################################################
